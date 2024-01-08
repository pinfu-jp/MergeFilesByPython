import os
import csv
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import PatternFill
import win32com.client

from module.logger import write_log, LogLevel
from module.datetime_util import get_datetime_by_str, get_yyyymmdd_by_datetime

COL_NO_TIMESTAMP = 1
COL_NO_FILE_NAME = 2
COL_NO_KEYWORD = 3
COL_NO_LOG = 4


def convert_marged_csvs_to_xlsx(csv_folder: str, xlsx_file: str):
    """ログマージしたCSVファイル群を1本のxlsxにまとめる"""

    write_log(f"convert_csvs_to_xlsx csv folder:{csv_folder} to xlsx:{xlsx_file}")

    # CSVフォルダ内のすべてのCSVファイルを取得する
    csv_pathes = [os.path.join(csv_folder, f) for f in os.listdir(csv_folder) if f.endswith('.csv')]
    csv_pathes = sorted(csv_pathes, reverse=True)

    # XLSXファイルを作成し、すべてのCSVファイルの内容を書き込む
    wb = openpyxl.Workbook()

    caution_color = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for csv_path in csv_pathes:

        # CSVを文字コードを加味して読み取り
        data, used_encoding = try_read_csv(csv_path)

        timestamp = get_datetime_by_str(os.path.splitext(os.path.basename(csv_path))[0])
        sheet_title = get_yyyymmdd_by_datetime(timestamp)

        write_log(f"create_sheet :{sheet_title}")
        ws = wb.create_sheet(title=sheet_title)
        
        # 行
        for row_no, row in enumerate(data, start=1):
            # 列
            for col_no, cell_value in enumerate(row, start=1):
                cell = ws.cell(row=row_no, column=col_no)
                cell.value = cell_value

                if row_no > 1 and col_no == COL_NO_KEYWORD:
                    if len(cell_value) > 0:
                        cell.fill = caution_color    # キーワードは注意色にする

        # カラム幅を調整
        # どのシートも同じ幅としたいので、固定値です
        ws.column_dimensions[get_column_letter(COL_NO_TIMESTAMP)].width = 26
        ws.column_dimensions[get_column_letter(COL_NO_FILE_NAME)].width = 64
        ws.column_dimensions[get_column_letter(COL_NO_KEYWORD)].width = 12

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet']) # 初期登録されている"sheet"は不要なので削除
    
    # ファイルを保存
    wb.save(xlsx_file)
    write_log(f"convert_csvs_to_xlsx end :{xlsx_file}")


def get_excel_ver():
    """Excelのバージョン取得"""
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        ver = float(excel.Version)
        excel.Quit()
        write_log(f"get_excel_ver :{ver}")
        return ver
    except:
        write_log("Excel is not installed.")
        return 0


def run_excel(file_path):
    """Excelを起動し、選択ファイルを開く"""

    try:
        write_log(f"Run Excel path:{file_path}")
        import win32com.client
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True
        workbook = excel.Workbooks.Open(file_path)
    except:
        write_log("Excel is not installed.")
        return


def try_read_csv(csv_path, encodings=('utf-8', 'cp932', 'iso-8859-1')):
    """指定されたエンコーディングを順に試してCSVファイルを読み込む"""

    for encoding in encodings:
        try:
            with open(csv_path, 'r', encoding=encoding) as f:
                reader = csv.reader(f)
                data = [row for row in reader]  # ファイルの内容を読み込む
            return data, encoding  # 成功したらデータと使用したエンコーディングを返す
        except UnicodeDecodeError:
            write_log("decode error path:" + csv_path + " encoding:" + encoding, LogLevel.W)
            continue  # このエンコーディングでは読み込めないので次を試す

    raise ValueError(f"Failed to read the file with provided encodings: {csv_path}")

def remove_formulas(text):
    """テキストからExcel数式を排除"""

    # 数式と思われるセル（等号で始まる）を変更
    if text.startswith('='):
        # ここで数式を除去するか、何らかの別の値に置き換える
        new_text = ''  # 数式を空文字で置き換え
    else:
        new_text = text

    return new_text


